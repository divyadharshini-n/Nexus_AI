"""
Labels Excel Exporter
Export global and local labels to Excel format
"""
from io import BytesIO
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class LabelsExcelExporter:
    """Export labels to Excel with formatting"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_labels(self, global_labels: List[Dict], local_labels: List[Dict], 
                     stage_name: str = "Stage") -> BytesIO:
        """
        Export global and local labels to Excel
        
        Args:
            global_labels: List of global label dictionaries
            local_labels: List of local label dictionaries
            stage_name: Name of the stage for the filename
            
        Returns:
            BytesIO: Excel file in memory
        """
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # Create Global Labels sheet
        self._create_global_labels_sheet(workbook, global_labels, stage_name)
        
        # Create Local Labels sheet
        self._create_local_labels_sheet(workbook, local_labels, stage_name)
        
        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output
    
    def _create_global_labels_sheet(self, workbook, global_labels: List[Dict], stage_name: str):
        """Create and format global labels sheet"""
        sheet = workbook.create_sheet("Global Labels")
        
        # Title
        sheet.merge_cells('A1:H1')
        title_cell = sheet['A1']
        title_cell.value = f"{stage_name} - Global Labels"
        title_cell.font = Font(bold=True, size=14, color="1F4E78")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers
        headers = [
            'Label Name',
            'Data Type',
            'Class',
            'Device Name',
            'Initial Value',
            'Constant',
            'Comment',
            'Remark'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data rows
        for row_num, label in enumerate(global_labels, 4):
            data = [
                label.get('name', ''),
                label.get('data_type', ''),
                label.get('class', ''),
                label.get('device', ''),
                label.get('initial_value', ''),
                label.get('constant', ''),
                label.get('comment', ''),
                label.get('comment', '')  # Remark (using comment for both)
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = self.border
        
        # Auto-adjust column widths
        for col_num in range(1, 9):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            for row in sheet[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        sheet.freeze_panes = 'A4'
    
    def _create_local_labels_sheet(self, workbook, local_labels: List[Dict], stage_name: str):
        """Create and format local labels sheet"""
        sheet = workbook.create_sheet("Local Labels")
        
        # Title
        sheet.merge_cells('A1:F1')
        title_cell = sheet['A1']
        title_cell.value = f"{stage_name} - Local Labels"
        title_cell.font = Font(bold=True, size=14, color="1F4E78")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers
        headers = [
            'Label Name',
            'Data Type',
            'Class',
            'Initial Value',
            'Constant',
            'Comment'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data rows
        for row_num, label in enumerate(local_labels, 4):
            data = [
                label.get('name', ''),
                label.get('data_type', ''),
                label.get('class', ''),
                label.get('initial_value', ''),
                label.get('constant', ''),
                label.get('comment', '')
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = self.border
        
        # Auto-adjust column widths
        for col_num in range(1, 7):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            for row in sheet[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        sheet.freeze_panes = 'A4'
    
    def export_all_stages_labels(self, stages_data: List[Dict]) -> BytesIO:
        """
        Export labels from all stages to a single Excel file
        
        Args:
            stages_data: List of dictionaries containing stage info and labels
                        Each dict should have: stage_number, stage_name, global_labels, local_labels
                        
        Returns:
            BytesIO: Excel file in memory
        """
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # Create sheets for each stage
        for stage_data in stages_data:
            stage_num = stage_data.get('stage_number', 0)
            stage_name = stage_data.get('stage_name', 'Unknown')
            global_labels = stage_data.get('global_labels', [])
            local_labels = stage_data.get('local_labels', [])
            
            # Create global labels sheet for this stage
            sheet_name = f"Stage {stage_num} - Global"
            if len(sheet_name) > 31:  # Excel sheet name limit
                sheet_name = f"S{stage_num}-Global"
            self._create_stage_global_sheet(workbook, global_labels, stage_num, stage_name, sheet_name)
            
            # Create local labels sheet for this stage
            sheet_name = f"Stage {stage_num} - Local"
            if len(sheet_name) > 31:
                sheet_name = f"S{stage_num}-Local"
            self._create_stage_local_sheet(workbook, local_labels, stage_num, stage_name, sheet_name)
        
        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output
    
    def _create_stage_global_sheet(self, workbook, global_labels: List[Dict], 
                                   stage_num: int, stage_name: str, sheet_name: str):
        """Create global labels sheet for a specific stage"""
        sheet = workbook.create_sheet(sheet_name)
        
        # Title
        sheet.merge_cells('A1:H1')
        title_cell = sheet['A1']
        title_cell.value = f"Stage {stage_num}: {stage_name} - Global Labels"
        title_cell.font = Font(bold=True, size=14, color="1F4E78")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers
        headers = [
            'Label Name', 'Data Type', 'Class', 'Device Name',
            'Initial Value', 'Constant', 'Comment', 'Remark'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data rows
        for row_num, label in enumerate(global_labels, 4):
            data = [
                label.get('name', ''),
                label.get('data_type', ''),
                label.get('class', ''),
                label.get('device', ''),
                label.get('initial_value', ''),
                label.get('constant', ''),
                label.get('comment', ''),
                label.get('comment', '')
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = self.border
        
        # Auto-adjust column widths
        for col_num in range(1, 9):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row in sheet[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        sheet.freeze_panes = 'A4'
    
    def _create_stage_local_sheet(self, workbook, local_labels: List[Dict],
                                  stage_num: int, stage_name: str, sheet_name: str):
        """Create local labels sheet for a specific stage"""
        sheet = workbook.create_sheet(sheet_name)
        
        # Title
        sheet.merge_cells('A1:F1')
        title_cell = sheet['A1']
        title_cell.value = f"Stage {stage_num}: {stage_name} - Local Labels"
        title_cell.font = Font(bold=True, size=14, color="1F4E78")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers
        headers = [
            'Label Name', 'Data Type', 'Class',
            'Initial Value', 'Constant', 'Comment'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data rows
        for row_num, label in enumerate(local_labels, 4):
            data = [
                label.get('name', ''),
                label.get('data_type', ''),
                label.get('class', ''),
                label.get('initial_value', ''),
                label.get('constant', ''),
                label.get('comment', '')
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = self.border
        
        # Auto-adjust column widths
        for col_num in range(1, 7):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row in sheet[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        sheet.freeze_panes = 'A4'


# Singleton instance
labels_excel_exporter = LabelsExcelExporter()
